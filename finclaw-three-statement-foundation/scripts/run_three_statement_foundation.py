from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


QUARTERS = ["Q1", "Q2", "Q3", "Q4"]
QUARTER_END_DATES = {"Q1": "2025-03-31", "Q2": "2025-06-30", "Q3": "2025-09-30", "Q4": "2025-12-31"}


PROFIT_CANONICAL = {
    "一、营业收入": "revenue",
    "减：营业成本": "operating_cost",
    "税金及附加": "tax_surcharge",
    "销售费用": "selling_expense",
    "广告费和业务宣传费": "advertising_expense",
    "管理费用": "management_expense",
    "业务招待费": "business_entertainment",
    "研究费用": "research_expense",
    "财务费用": "finance_expense",
    "其中：利息费用（收入以“-”号填列）": "interest_expense",
    "加：投资收益（损失以“-”号填列）": "investment_income",
    "二、营业利润（亏损以“-”号填列）": "operating_profit",
    "加：营业外收入": "non_operating_income",
    "减：营业外支出": "non_operating_expense",
    "三、利润总额（亏损总额以“-”号填列）": "profit_total",
    "减：所得税费用": "income_tax",
    "四：净利润（净亏损以“-”号填列）": "net_profit",
}


CASH_CANONICAL = {
    "销售产成品、商品、提供劳务收到的现金": "sales_cash_received",
    "收到其他与经营活动有关的现金": "other_operating_cash_received",
    "购买原材料、商品、接受劳务支付的现金": "purchase_cash_paid",
    "支付的职工薪酬": "salary_cash_paid",
    "支付的税费": "tax_cash_paid",
    "支付其他与经营活动有关的现金": "other_operating_cash_paid",
    "经营活动产生的现金流量净额": "operating_cash_flow",
    "购建固定资产、无形资产和其他非流动资产支付的现金": "asset_purchase_cash_paid",
    "投资活动产生的现金流量净额": "investing_cash_flow",
    "取得借款收到的现金": "borrowing_cash_received",
    "吸收投资者投资收到的现金": "capital_injection_cash_received",
    "偿还借款本金支付的现金": "borrowing_principal_paid",
    "偿还借款利息支付的现金": "borrowing_interest_paid",
    "分配利润支付的现金": "dividend_paid",
    "筹资活动产生的现金流量净额": "financing_cash_flow",
    "四、现金净增加额": "net_cash_increase",
    "加：期初现金余额": "beginning_cash",
    "五、期末现金余额": "ending_cash",
}


BALANCE_CANONICAL = {
    "货币资金": "cash",
    "应收账款": "accounts_receivable",
    "预付账款": "prepayments",
    "其他应收款": "other_receivables",
    "其他流动资产": "other_current_assets",
    "流动资产合计": "current_assets_total",
    "固定资产原价": "fixed_assets_original",
    "减：累计折旧": "accumulated_depreciation",
    "固定资产账面价值": "fixed_assets_book",
    "无形资产": "intangible_assets",
    "其他非流动资产": "other_non_current_assets",
    "非流动资产合计": "non_current_assets_total",
    "资产总计": "assets_total",
    "短期借款": "short_term_borrowing",
    "应付账款": "accounts_payable",
    "预收账款": "advance_receipts",
    "应付职工薪酬": "salary_payable",
    "应交税费": "taxes_payable",
    "其他应付款": "other_payables",
    "其他流动负债": "other_current_liabilities",
    "流动负债合计：": "current_liabilities_total",
    "长期借款": "long_term_borrowing",
    "递延收益": "deferred_income",
    "非流动负债合计": "non_current_liabilities_total",
    "负债合计": "liabilities_total",
    "实收资本（或股本）": "paid_in_capital",
    "资本公积": "capital_reserve",
    "盈余公积": "surplus_reserve",
    "未分配利润": "retained_earnings",
    "所有者权益（或股东权益）合计": "equity_total",
    "负债和所有者权益（或股东权益）总计": "liabilities_equity_total",
}


@dataclass
class SourceRef:
    file: str
    sheet: str
    row: int
    column: str


@dataclass
class Fact:
    company: str
    period: str
    period_type: str
    statement_type: str
    line_item_raw: str
    line_item_canonical: str
    amount_type: str
    amount: float | None
    unit: str
    source: SourceRef
    quality_flags: list[str]


def clean_item(value: Any) -> str:
    return str(value or "").strip()


def num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    try:
        return round(float(str(value).replace(",", "")), 2)
    except ValueError:
        return None


def file_quarter(path: Path) -> str | None:
    match = re.search(r"2025年(Q[1-4])", path.name)
    return match.group(1) if match else None


def detect_statement_type(path: Path, title: str) -> str | None:
    text = path.name + " " + title
    if "利润表" in text:
        return "profit_statement"
    if "资产负债表" in text:
        return "balance_sheet"
    if "现金流量表" in text:
        return "cash_flow_statement"
    return None


def parse_company(value: Any) -> str:
    text = clean_item(value)
    return text.split("编制单位：", 1)[1].strip() if "编制单位：" in text else text


def parse_unit(row_values: list[Any]) -> str:
    for value in row_values:
        text = clean_item(value)
        if text.startswith("单位："):
            return text.split("单位：", 1)[1].strip()
    return ""


def issue(severity: str, issue_type: str, message: str, **extra: Any) -> dict[str, Any]:
    return {"severity": severity, "issue_type": issue_type, "message": message, **extra}


def add_fact(facts: list[Fact], meta: dict[str, Any], period_type: str, line_raw: str, line_canonical: str, amount_type: str, amount: float | None, row: int, column: str, flags: list[str] | None = None) -> None:
    facts.append(
        Fact(
            company=meta["company"],
            period=meta["period"],
            period_type=period_type,
            statement_type=meta["statement_type"],
            line_item_raw=line_raw,
            line_item_canonical=line_canonical,
            amount_type=amount_type,
            amount=amount,
            unit=meta["unit"],
            source=SourceRef(file=meta["file"], sheet=meta["sheet"], row=row, column=column),
            quality_flags=flags or [],
        )
    )


def parse_workbook(path: Path, expected_year: int) -> tuple[list[Fact], dict[str, Any], list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    title = ws.title
    row2 = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    statement_type = detect_statement_type(path, title)
    period = file_quarter(path)
    company = parse_company(ws["A2"].value)
    unit = parse_unit(row2)
    issues: list[dict[str, Any]] = []
    facts: list[Fact] = []
    if not statement_type:
        issues.append(issue("blocking_error", "statement_type_unknown", f"无法识别报表类型: {path.name}", file=path.name))
        return facts, {"file": path.name, "sheet": title, "company": company, "period": period, "unit": unit, "statement_type": None}, issues
    if not period:
        issues.append(issue("blocking_error", "period_unknown", f"无法从文件名识别季度: {path.name}", file=path.name))
    meta = {"file": path.name, "sheet": title, "company": company, "period": period or "UNKNOWN", "unit": unit, "statement_type": statement_type}
    if statement_type == "balance_sheet":
        date = clean_item(ws["E2"].value)
        expected_date = QUARTER_END_DATES.get(period or "")
        if expected_date and date != expected_date.replace("2025", str(expected_year)):
            issues.append(issue("blocking_error", "period_mismatch", f"{path.name} 表内日期 {date} 与文件季度 {period} 不一致", file=path.name, expected=expected_date, actual=date))
        for idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            left_item = clean_item(row[0])
            if left_item:
                canonical = BALANCE_CANONICAL.get(left_item, left_item)
                add_fact(facts, meta, "point_in_time", left_item, canonical, "ending", num(row[2]), idx, "期末余额")
                add_fact(facts, meta, "point_in_time", left_item, canonical, "beginning", num(row[3]), idx, "年初余额")
            right_item = clean_item(row[4]) if len(row) > 4 else ""
            if right_item:
                canonical = BALANCE_CANONICAL.get(right_item, right_item)
                add_fact(facts, meta, "point_in_time", right_item, canonical, "ending", num(row[6]), idx, "期末余额")
                add_fact(facts, meta, "point_in_time", right_item, canonical, "beginning", num(row[7]), idx, "年初余额")
    elif statement_type == "profit_statement":
        for idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            item = clean_item(row[0])
            if not item:
                continue
            canonical = PROFIT_CANONICAL.get(item, item)
            add_fact(facts, meta, "year_to_date", item, canonical, "ytd", num(row[2]), idx, "本年累计金额")
            add_fact(facts, meta, "current_period", item, canonical, "current", num(row[3]), idx, "本期金额")
    elif statement_type == "cash_flow_statement":
        for idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            item = clean_item(row[0])
            if not item:
                continue
            canonical = CASH_CANONICAL.get(item, item)
            add_fact(facts, meta, "current_period", item, canonical, "current", num(row[2]), idx, "本期金额")
            add_fact(facts, meta, "year_to_date", item, canonical, "ytd", num(row[3]), idx, "本年累计金额")
    return facts, meta, issues


def fact_lookup(facts: list[Fact]) -> dict[tuple[str, str, str, str], float | None]:
    return {(f.period, f.statement_type, f.line_item_canonical, f.amount_type): f.amount for f in facts}


def value(lookup: dict[tuple[str, str, str, str], float | None], period: str, statement_type: str, item: str, amount_type: str) -> float | None:
    return lookup.get((period, statement_type, item, amount_type))


def validate_collection(facts: list[Fact], metas: list[dict[str, Any]], expected_year: int, expected_company: str | None) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    companies = sorted({m.get("company") for m in metas if m.get("company")})
    if expected_company and companies != [expected_company]:
        issues.append(issue("blocking_error", "company_mismatch", "文件公司名与期望公司不一致", expected=expected_company, actual=companies))
    elif len(companies) > 1:
        issues.append(issue("blocking_error", "multiple_companies", "输入文件包含多个公司，不能作为单体报表分析", actual=companies))
    units = sorted({m.get("unit") for m in metas if m.get("unit")})
    if len(units) > 1:
        issues.append(issue("warning", "unit_mismatch", "输入文件单位不完全一致", actual=units))
    expected_pairs = {(q, st) for q in QUARTERS for st in ["profit_statement", "balance_sheet", "cash_flow_statement"]}
    actual_pairs = {(m.get("period"), m.get("statement_type")) for m in metas}
    for period, statement_type in sorted(expected_pairs - actual_pairs):
        issues.append(issue("blocking_error", "missing_statement", f"缺少 {period} {statement_type}", period=period, statement_type=statement_type))
    lookup = fact_lookup(facts)
    for q in QUARTERS:
        assets = value(lookup, q, "balance_sheet", "assets_total", "ending")
        liabilities_equity = value(lookup, q, "balance_sheet", "liabilities_equity_total", "ending")
        if assets is not None and liabilities_equity is not None and abs(assets - liabilities_equity) > 0.01:
            issues.append(issue("blocking_error", "balance_not_balanced", f"{q} 资产负债表不平衡", period=q, assets_total=assets, liabilities_equity_total=liabilities_equity))
        bs_cash = value(lookup, q, "balance_sheet", "cash", "ending")
        cf_cash = value(lookup, q, "cash_flow_statement", "ending_cash", "current")
        if bs_cash is not None and cf_cash is not None and abs(bs_cash - cf_cash) > 0.01:
            issues.append(issue("blocking_error", "cash_tieout_failed", f"{q} 现金流期末现金与资产负债表货币资金不一致", period=q, balance_cash=bs_cash, cash_flow_ending_cash=cf_cash))
    return issues


def metrics_by_quarter(facts: list[Fact]) -> dict[str, dict[str, float | None]]:
    lookup = fact_lookup(facts)
    metrics: dict[str, dict[str, float | None]] = {}
    for q in QUARTERS:
        revenue = value(lookup, q, "profit_statement", "revenue", "current")
        net_profit = value(lookup, q, "profit_statement", "net_profit", "current")
        operating_cash_flow = value(lookup, q, "cash_flow_statement", "operating_cash_flow", "current")
        cash = value(lookup, q, "balance_sheet", "cash", "ending")
        assets = value(lookup, q, "balance_sheet", "assets_total", "ending")
        liabilities = value(lookup, q, "balance_sheet", "liabilities_total", "ending")
        equity = value(lookup, q, "balance_sheet", "equity_total", "ending")
        metrics[q] = {
            "revenue": revenue,
            "management_expense": value(lookup, q, "profit_statement", "management_expense", "current"),
            "research_expense": value(lookup, q, "profit_statement", "research_expense", "current"),
            "operating_profit": value(lookup, q, "profit_statement", "operating_profit", "current"),
            "net_profit": net_profit,
            "net_profit_ytd": value(lookup, q, "profit_statement", "net_profit", "ytd"),
            "operating_cash_flow": operating_cash_flow,
            "ending_cash": cash,
            "assets_total": assets,
            "liabilities_total": liabilities,
            "equity_total": equity,
            "net_margin": round(net_profit / revenue, 4) if revenue not in (None, 0) and net_profit is not None else None,
            "ocf_to_net_profit": round(operating_cash_flow / net_profit, 4) if net_profit not in (None, 0) and operating_cash_flow is not None else None,
            "asset_liability_ratio": round(liabilities / assets, 4) if assets not in (None, 0) and liabilities is not None else None,
        }
    return metrics


def pct_change(current: float | None, previous: float | None) -> float | None:
    if current is None or previous in (None, 0):
        return None
    return round((current - previous) / abs(previous), 4)


def build_analysis(metrics: dict[str, dict[str, float | None]], issues: list[dict[str, Any]]) -> dict[str, Any]:
    q4 = metrics["Q4"]
    q3 = metrics["Q3"]
    lowest_profit_q = min(QUARTERS, key=lambda q: metrics[q]["net_profit"] if metrics[q]["net_profit"] is not None else float("inf"))
    highest_cash_q = max(QUARTERS, key=lambda q: metrics[q]["ending_cash"] if metrics[q]["ending_cash"] is not None else float("-inf"))
    return {
        "has_blocking_errors": any(i["severity"] == "blocking_error" for i in issues),
        "profit_trend": {
            "lowest_net_profit_quarter": lowest_profit_q,
            "q4_net_profit": q4["net_profit"],
            "q4_net_profit_qoq_change": pct_change(q4["net_profit"], q3["net_profit"]),
            "q4_revenue_qoq_change": pct_change(q4["revenue"], q3["revenue"]),
        },
        "cash_quality": {
            "q4_ocf_to_net_profit": q4["ocf_to_net_profit"],
            "highest_cash_quarter": highest_cash_q,
        },
        "solvency": {
            q: {
                "asset_liability_ratio": metrics[q]["asset_liability_ratio"],
                "equity_total": metrics[q]["equity_total"],
            }
            for q in QUARTERS
        },
    }


def fmt(value_: float | None) -> str:
    if value_ is None:
        return "N/A"
    return f"{value_:,.2f}"


def build_report(company: str, metrics: dict[str, dict[str, float | None]], analysis: dict[str, Any], issues: list[dict[str, Any]]) -> str:
    if analysis["has_blocking_errors"]:
        blocking = [i for i in issues if i["severity"] == "blocking_error"]
        lines = [f"# {company} 三张表分析数据校验未通过", "", "发现阻断问题，暂不生成正式财务分析报告。", ""]
        for item in blocking:
            lines.append(f"- {item['issue_type']}: {item['message']}")
        return "\n".join(lines) + "\n"
    q4 = metrics["Q4"]
    lines = [
        f"# {company} 2025 年 Q1-Q4 三张表分析草稿",
        "",
        "## 核心结论",
        f"- Q4 营业收入为 {fmt(q4['revenue'])} 元，净利润为 {fmt(q4['net_profit'])} 元。",
        f"- Q4 经营活动现金流为 {fmt(q4['operating_cash_flow'])} 元，期末货币资金为 {fmt(q4['ending_cash'])} 元。",
        f"- Q4 资产负债率为 {q4['asset_liability_ratio'] if q4['asset_liability_ratio'] is not None else 'N/A'}。",
        "",
        "## 季度关键指标",
        "| 指标 | Q1 | Q2 | Q3 | Q4 |",
        "| --- | ---: | ---: | ---: | ---: |",
    ]
    for label, key in [
        ("营业收入", "revenue"),
        ("管理费用", "management_expense"),
        ("研发费用", "research_expense"),
        ("营业利润", "operating_profit"),
        ("净利润", "net_profit"),
        ("经营现金流", "operating_cash_flow"),
        ("期末货币资金", "ending_cash"),
        ("资产总计", "assets_total"),
        ("负债合计", "liabilities_total"),
    ]:
        lines.append("| " + label + " | " + " | ".join(fmt(metrics[q][key]) for q in QUARTERS) + " |")
    lines.extend([
        "",
        "## 数据边界",
        "- 当前报告基于财务三张表生成，未使用经营数据。",
        "- 订阅收入、广告收入、用户指标、人效指标等经营分析需要后续接入业务数据后再补齐。",
    ])
    warning_items = [i for i in issues if i["severity"] != "blocking_error"]
    if warning_items:
        lines.extend(["", "## 数据提示"])
        lines.extend([f"- {i['severity']}: {i['message']}" for i in warning_items])
    return "\n".join(lines) + "\n"


def build_followups(metrics: dict[str, dict[str, float | None]]) -> list[dict[str, Any]]:
    return [
        {
            "question": "Q4 利润表现如何？",
            "answer": f"Q4 净利润为 {fmt(metrics['Q4']['net_profit'])} 元，营业收入为 {fmt(metrics['Q4']['revenue'])} 元，净利率为 {metrics['Q4']['net_margin'] if metrics['Q4']['net_margin'] is not None else 'N/A'}。",
            "evidence": ["profit_statement.net_profit.current.Q4", "profit_statement.revenue.current.Q4"],
        },
        {
            "question": "经营现金流和净利润是否匹配？",
            "answer": f"Q4 经营现金流为 {fmt(metrics['Q4']['operating_cash_flow'])} 元，净利润为 {fmt(metrics['Q4']['net_profit'])} 元，经营现金流/净利润为 {metrics['Q4']['ocf_to_net_profit'] if metrics['Q4']['ocf_to_net_profit'] is not None else 'N/A'}。",
            "evidence": ["cash_flow_statement.operating_cash_flow.current.Q4", "profit_statement.net_profit.current.Q4"],
        },
        {
            "question": "没有业务数据时能分析什么？",
            "answer": "可以完成三张表趋势、利润质量、现金流质量、偿债能力和勾稽校验；用户、流量、广告、ARPPU、人效等经营归因需要补充业务数据。",
            "evidence": ["template_gap.business_data"],
        },
    ]


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_facts_csv(path: Path, facts: list[Fact]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "company",
                "period",
                "period_type",
                "statement_type",
                "line_item_raw",
                "line_item_canonical",
                "amount_type",
                "amount",
                "unit",
                "source_file",
                "source_sheet",
                "source_row",
                "source_column",
                "quality_flags",
            ],
        )
        writer.writeheader()
        for fact in facts:
            writer.writerow({
                "company": fact.company,
                "period": fact.period,
                "period_type": fact.period_type,
                "statement_type": fact.statement_type,
                "line_item_raw": fact.line_item_raw,
                "line_item_canonical": fact.line_item_canonical,
                "amount_type": fact.amount_type,
                "amount": fact.amount,
                "unit": fact.unit,
                "source_file": fact.source.file,
                "source_sheet": fact.source.sheet,
                "source_row": fact.source.row,
                "source_column": fact.source.column,
                "quality_flags": ";".join(fact.quality_flags),
            })


def run_backtest(metrics: dict[str, dict[str, float | None]], expected_path: Path | None) -> dict[str, Any] | None:
    if not expected_path:
        return None
    expected = json.loads(expected_path.read_text(encoding="utf-8"))
    checks = []
    mapping = {
        "revenue_current": "revenue",
        "management_expense_current": "management_expense",
        "operating_profit_current": "operating_profit",
        "net_profit_current": "net_profit",
        "net_profit_ytd": "net_profit_ytd",
        "cash_ending_balance_sheet": "ending_cash",
        "cash_ending_cash_flow": "ending_cash",
        "operating_cash_flow_current": "operating_cash_flow",
        "assets_total": "assets_total",
        "liabilities_equity_total": "assets_total",
    }
    for q, expected_metrics in expected["key_metrics"].items():
        for expected_key, metric_key in mapping.items():
            actual = metrics[q].get(metric_key)
            expected_value = expected_metrics.get(expected_key)
            passed = actual is not None and expected_value is not None and abs(float(actual) - float(expected_value)) <= 0.01
            checks.append({"period": q, "expected_key": expected_key, "metric_key": metric_key, "actual": actual, "expected": expected_value, "passed": passed})
    return {"passed": all(c["passed"] for c in checks), "checks": checks}


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse, validate, and analyze Kingdee-like Q1-Q4 financial statements for FinClaw.")
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--expected-year", type=int, default=2025)
    parser.add_argument("--expected-company")
    parser.add_argument("--expected-results")
    args = parser.parse_args()

    input_dir = Path(args.input_dir).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    expected_results = Path(args.expected_results).expanduser().resolve() if args.expected_results else None

    files = sorted(input_dir.glob("*.xlsx"))
    all_facts: list[Fact] = []
    metas: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    for path in files:
        facts, meta, file_issues = parse_workbook(path, args.expected_year)
        all_facts.extend(facts)
        metas.append(meta)
        issues.extend(file_issues)
    issues.extend(validate_collection(all_facts, metas, args.expected_year, args.expected_company))
    metrics = metrics_by_quarter(all_facts)
    analysis = build_analysis(metrics, issues)
    company = args.expected_company or (metas[0]["company"] if metas else "UNKNOWN")
    report = build_report(company, metrics, analysis, issues)
    followups = build_followups(metrics)
    backtest = run_backtest(metrics, expected_results)

    facts_dict = []
    for fact in all_facts:
        row = asdict(fact)
        facts_dict.append(row)
    outputs = {
        "validation_issues": str(out_dir / "validation_issues.json"),
        "statement_facts_json": str(out_dir / "statement_facts.json"),
        "statement_facts_csv": str(out_dir / "statement_facts.csv"),
        "quarter_metrics": str(out_dir / "quarter_metrics.json"),
        "analysis_summary": str(out_dir / "analysis_summary.json"),
        "report_draft": str(out_dir / "report_draft.md"),
        "followup_examples": str(out_dir / "followup_examples.json"),
    }
    write_json(out_dir / "validation_issues.json", issues)
    write_json(out_dir / "statement_facts.json", facts_dict)
    write_facts_csv(out_dir / "statement_facts.csv", all_facts)
    write_json(out_dir / "quarter_metrics.json", metrics)
    write_json(out_dir / "analysis_summary.json", analysis)
    (out_dir / "report_draft.md").write_text(report, encoding="utf-8")
    write_json(out_dir / "followup_examples.json", followups)
    if backtest is not None:
        write_json(out_dir / "backtest_results.json", backtest)
        outputs["backtest_results"] = str(out_dir / "backtest_results.json")
    manifest = {
        "input_dir": str(input_dir),
        "file_count": len(files),
        "files": [m for m in metas],
        "has_blocking_errors": analysis["has_blocking_errors"],
        "outputs": outputs,
    }
    write_json(out_dir / "manifest.json", manifest)
    print(json.dumps({"out_dir": str(out_dir), "has_blocking_errors": analysis["has_blocking_errors"], "issue_count": len(issues), "outputs": outputs}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
