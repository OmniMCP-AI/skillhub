"""Write a reviewable boss-facing workbook from normalized report objects."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, List

from .consistency_checks import ValidationIssue
from .dashboard_rows import DashboardRow
from .formatting import fmt_ratio, fmt_yuan_as_wan
from .metrics import MetricStore
from .narrative_values import build_management_summary
from .statement_facts import QUARTERS, ReportFacts


def write_boss_review_workbook(
    output_path: str | Path,
    *,
    facts: ReportFacts,
    metrics: MetricStore,
    dashboard_rows: List[DashboardRow],
    validation_issues: Iterable[ValidationIssue] = (),
) -> Path:
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill
    except ModuleNotFoundError as exc:
        raise RuntimeError("Missing dependency openpyxl. Install it in the Hermes runtime/sandbox that writes workbooks.") from exc

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def sheet(name: str, headers: List[str]):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F5597")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        return ws

    cover = sheet("封面", ["项目", "内容", "辅助项目", "辅助内容"])
    summaries = build_management_summary(metrics)
    cover_rows = [
        ["公司", facts.company or "未识别", "报告期间", f"{facts.year or ''} Q1-Q4"],
        ["报告名称", f"{facts.company or '公司'} 财报与经营分析", "报告模板", "FinClaw 老板审阅"],
        ["数据口径", "三表标准口径：利润表/现金流量表按单季与累计，资产负债表按期末", "金额单位", "人民币万元"],
        ["核心结论", summaries[0], "现金质量", summaries[1]],
        ["限制事项", "本输出由稳定脚本生成，仍需人工复核经营明细、预算、行业基准与披露边界", "数据性质", "FIN_STMT / SYNTHETIC_DEMO 由调用方指定"],
    ]
    _append_rows(cover, cover_rows)

    boss = sheet("老板摘要", ["模块", "管理层结论（事实+数字）", "决策提示（管理含义）"])
    boss_rows = [
        ["总体表现", summaries[0], "看收入、利润、现金是否同向增长。"],
        ["增长质量", summaries[2], "环比必须基于单季收入，不能用累计收入代替。"],
        ["现金质量", summaries[1], "同时看累计和单季 OCF/净利润，避免口径混淆。"],
        ["偿债结构", summaries[3], "期末指标使用资产负债表期末数。"],
        ["管理动作", "跟踪应收账款账龄、研发投入产出、现金使用规划。", "下一版可补客户/产品/订单/预算数据。"],
    ]
    _append_rows(boss, boss_rows)

    overview = sheet("经营概览", ["模块", "事实与数据", "管理观察"])
    _append_rows(
        overview,
        [
            ["规模", summaries[0], "收入和利润使用累计口径。"],
            ["季度增长", summaries[2], "环比使用单季收入。"],
            ["现金", summaries[1], "现金质量同时展示单季与累计口径。"],
            ["资产负债", summaries[3], "资产、负债、权益均为期末数。"],
        ],
    )

    _write_metric_table(sheet("利润分析", ["指标", "Q1", "Q2", "Q3", "Q4", "全年/期末", "管理观察"]), metrics, [
        ("营业收入（万元，累计）", "revenue.ytd", "累计口径"),
        ("营业收入（单季，万元）", "revenue.current", "单季口径"),
        ("毛利率（累计）", "gross_margin.ytd", "累计口径"),
        ("净利率（累计）", "net_margin.ytd", "累计口径"),
        ("净利润（万元，累计）", "net_profit.ytd", "累计口径"),
    ])
    _write_metric_table(sheet("资产负债分析", ["指标", "Q1", "Q2", "Q3", "Q4", "全年/期末", "管理观察"]), metrics, [
        ("期末货币资金（万元）", "cash.bs.ending", "期末数"),
        ("应收账款（万元，期末）", "accounts_receivable.ending", "期末数"),
        ("资产总计（万元）", "total_assets.ending", "期末数"),
        ("资产负债率（期末）", "debt_to_asset.ending", "期末数"),
    ])
    _write_metric_table(sheet("现金流分析", ["指标", "Q1", "Q2", "Q3", "Q4", "全年/期末", "管理观察"]), metrics, [
        ("经营现金流（累计，万元）", "ocf.ytd", "累计口径"),
        ("OCF / 净利润（单季）", "ocf_net_profit_ratio.current", "单季口径"),
        ("OCF / 净利润（累计）", "ocf_net_profit_ratio.ytd", "累计口径"),
    ])

    key_sheet = sheet("关键指标", ["指标", "Q1", "Q2", "Q3", "Q4", "全年/期末", "口径 / 说明"])
    dashboard = sheet("财务经营Dashboard", ["指标", "Q1", "Q2", "Q3", "Q4", "全年/期末", "口径 / 说明"])
    for ws in (key_sheet, dashboard):
        for row in dashboard_rows:
            ws.append([row.label, row.values["Q1"], row.values["Q2"], row.values["Q3"], row.values["Q4"], row.full_year, row.note])

    risk = sheet("风险与建议", ["风险/机会", "事实依据", "判断", "建议"])
    _append_rows(
        risk,
        [
            ["口径风险", "单季/累计/期末指标已显式拆分", "可控", "后续 writer 必须复用 normalized metrics"],
            ["现金质量", summaries[1], "需持续跟踪", "保留 OCF/净利润单季与累计两行"],
            ["经营明细缺口", "客户/产品/订单/账龄未在三表中提供", "影响经营驱动解释", "补充经营数据后再做驱动拆解"],
        ],
    )

    support = sheet("追问支持", ["用户可能追问", "当前可回答范围", "需要补充的数据", "建议回答口径"])
    _append_rows(
        support,
        [
            ["为什么环比不等于累计增长？", "可回答", "无", "环比只用单季收入计算"],
            ["现金质量为什么好？", "可回答", "无", "使用 OCF/净利润单季和累计指标"],
            ["增长来自哪些客户？", "不可完全回答", "客户/合同/订单明细", "说明三表不能穿透客户结构"],
        ],
    )

    audit = sheet("数据与计算审核报告", ["检查项", "结论", "说明"])
    issue_rows = [[issue.code, issue.severity, issue.message] for issue in validation_issues]
    _append_rows(audit, issue_rows or [["一致性校验", "PASS", "未发现阻断问题"]])

    lineage = sheet("关键数据溯源摘要", ["指标", "季度", "口径", "单位", "公式", "来源"])
    for point in sorted(metrics.all(), key=lambda p: (p.key, p.quarter)):
        lineage.append([point.label, point.quarter, point.grain, point.unit, point.formula, "; ".join(point.source_refs)])

    limit = sheet("限制事项与下一步", ["事项", "说明"])
    _append_rows(
        limit,
        [
            ["限制事项", "当前脚本只固化三表指标和口径一致性，不替代经营数据补充与人工审阅。"],
            ["下一步 1", "补客户/产品/订单/账龄，做经营驱动分析。"],
            ["下一步 2", "补预算与行业基准，做预算偏差和对标。"],
            ["下一步 3", "将脚本接入 Hermes runtime 前，先用多公司样本回归测试。"],
        ],
    )

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 24
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def _append_rows(ws, rows):
    for row in rows:
        ws.append(row)


def _write_metric_table(ws, metrics: MetricStore, rows):
    for label, key, note in rows:
        values = []
        for quarter in QUARTERS:
            point = metrics.get(key, quarter)
            if point is None:
                values.append("-")
            elif point.unit == "ratio":
                values.append(fmt_ratio(point.value))
            else:
                values.append(fmt_yuan_as_wan(point.value))
        full_year = values[-1] if not key.endswith(".current") else "-"
        ws.append([label, *values, full_year, note])
