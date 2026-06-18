"""Parse Chinese three-statement Excel files into raw statement facts.

The parser intentionally reads statement structure rather than hard-coding the
Shanghai Yunling demo files.  It supports the common FinClaw mock-data layouts:

- 利润表: 项目 / 行次 / 本年累计金额 / 本期金额
- 现金流量表: 项目 / 本期金额 / 本年累计金额
- 资产负债表: left asset block and right liability/equity block
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple

QUARTERS = ("Q1", "Q2", "Q3", "Q4")

StatementDict = Dict[str, Dict[str, Optional[float]]]


class StatementParseError(RuntimeError):
    """Raised when a required statement or period cannot be parsed."""


@dataclass(frozen=True)
class SourceCell:
    file: str
    sheet: str
    row: int
    column: int
    item: str
    field: str
    value: Optional[float]


@dataclass
class ReportFacts:
    company: str = ""
    year: Optional[int] = None
    quarters: Tuple[str, ...] = QUARTERS
    profit_current: StatementDict = field(default_factory=dict)
    profit_ytd: StatementDict = field(default_factory=dict)
    cashflow_current: StatementDict = field(default_factory=dict)
    cashflow_ytd: StatementDict = field(default_factory=dict)
    balance_assets_ending: StatementDict = field(default_factory=dict)
    balance_liability_equity_ending: StatementDict = field(default_factory=dict)
    lineage: Dict[str, SourceCell] = field(default_factory=dict)
    source_files: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    def lineage_key(self, statement: str, quarter: str, item: str, field: str) -> str:
        return f"{statement}.{quarter}.{item}.{field}"


def parse_statement_directory(input_dir: str | Path, *, company: str = "", year: Optional[int] = None) -> ReportFacts:
    """Parse all xlsx files under ``input_dir`` into ``ReportFacts``.

    ``openpyxl`` is imported lazily so agents can inspect and test most of the
    package without having Excel dependencies installed.
    """

    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError as exc:
        raise StatementParseError(
            "Missing dependency openpyxl. Install it in the Hermes runtime/sandbox "
            "that executes FinClaw report parsing."
        ) from exc

    root = Path(input_dir)
    if not root.exists():
        raise StatementParseError(f"Input directory does not exist: {root}")

    files = sorted([p for p in root.iterdir() if p.suffix.lower() in {".xlsx", ".xlsm"} and not p.name.startswith("~$")])
    if not files:
        raise StatementParseError(f"No Excel statement files found in: {root}")

    facts = ReportFacts(company=company, year=year)
    facts.source_files = [str(p) for p in files]

    for path in files:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        sheet_title = ws.title
        kind = _detect_statement_kind(path.name, sheet_title)
        quarter = _detect_quarter(path.name, sheet_title)
        if not kind or not quarter:
            facts.warnings.append(f"Skipped {path.name}: unable to detect statement kind or quarter")
            continue

        if not facts.year:
            facts.year = _detect_year(path.name, sheet_title)
        if not facts.company:
            facts.company = _detect_company(ws.cell(row=2, column=1).value) or company

        if kind == "profit":
            current, ytd, lineage = _parse_two_amount_statement(
                ws,
                path=path,
                statement="profit",
                quarter=quarter,
                current_header="本期金额",
                ytd_header="本年累计金额",
            )
            facts.profit_current[quarter] = current
            facts.profit_ytd[quarter] = ytd
            facts.lineage.update(lineage)
        elif kind == "cashflow":
            current, ytd, lineage = _parse_two_amount_statement(
                ws,
                path=path,
                statement="cashflow",
                quarter=quarter,
                current_header="本期金额",
                ytd_header="本年累计金额",
            )
            facts.cashflow_current[quarter] = current
            facts.cashflow_ytd[quarter] = ytd
            facts.lineage.update(lineage)
        elif kind == "balance":
            assets, liab_equity, lineage = _parse_balance_sheet(ws, path=path, quarter=quarter)
            facts.balance_assets_ending[quarter] = assets
            facts.balance_liability_equity_ending[quarter] = liab_equity
            facts.lineage.update(lineage)

    _validate_required_quarters(facts)
    return facts


def _detect_statement_kind(filename: str, sheet_title: str) -> Optional[str]:
    text = f"{filename} {sheet_title}"
    if "资产负债表" in text:
        return "balance"
    if "现金流量表" in text:
        return "cashflow"
    if "利润表" in text:
        return "profit"
    return None


def _detect_quarter(filename: str, sheet_title: str) -> Optional[str]:
    text = f"{filename} {sheet_title}"
    for idx, quarter in enumerate(QUARTERS, start=1):
        if re.search(rf"\bQ{idx}\b", text, flags=re.IGNORECASE):
            return quarter
        if f"第{idx}季度" in text or f"{idx}季度" in text:
            return quarter
    period_match = re.search(r"(\d{4})年([369]|12)期", text)
    if period_match:
        return {"3": "Q1", "6": "Q2", "9": "Q3", "12": "Q4"}[period_match.group(2)]
    date_match = re.search(r"\d{4}[-年/]([0369]{1,2}|12)[-/月]", text)
    if date_match:
        return {"03": "Q1", "3": "Q1", "06": "Q2", "6": "Q2", "09": "Q3", "9": "Q3", "12": "Q4"}.get(date_match.group(1))
    return None


def _detect_year(filename: str, sheet_title: str) -> Optional[int]:
    match = re.search(r"(20\d{2})", f"{filename} {sheet_title}")
    return int(match.group(1)) if match else None


def _detect_company(cell_value: Any) -> str:
    if not cell_value:
        return ""
    text = str(cell_value)
    text = text.replace("编制单位：", "").replace("编制单位:", "").strip()
    return text


def _normalize_item(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", "", text)
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("“", '"').replace("”", '"').replace("－", "-").replace("—", "-").replace("–", "-")
    aliases = {
        "研究费用": "研发费用",
        "四：净利润(净亏损以\"-\"号填列)": "四、净利润(净亏损以\"-\"号填列)",
        "所有者权益(或股东权益)合计": "所有者权益合计",
        "所有者权益（或股东权益）合计": "所有者权益合计",
        "负债和所有者权益(或股东权益)总计": "负债和所有者权益总计",
        "负债和所有者权益（或股东权益）总计": "负债和所有者权益总计",
    }
    return aliases.get(text, text)


def _as_number(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _find_header_row(rows: Iterable[Tuple[Any, ...]]) -> Tuple[int, Tuple[Any, ...]]:
    for idx, row in enumerate(rows, start=1):
        normalized = [_normalize_item(v) for v in row]
        if "项目" in normalized or "资产" in normalized:
            return idx, row
    raise StatementParseError("Unable to find statement header row")


def _column_index(header: Tuple[Any, ...], label: str) -> Optional[int]:
    normalized = [_normalize_item(v) for v in header]
    for idx, value in enumerate(normalized, start=1):
        if value == _normalize_item(label):
            return idx
    return None


def _parse_two_amount_statement(ws: Any, *, path: Path, statement: str, quarter: str, current_header: str, ytd_header: str) -> Tuple[Dict[str, Optional[float]], Dict[str, Optional[float]], Dict[str, SourceCell]]:
    rows = list(ws.iter_rows(values_only=True))
    header_row_idx, header = _find_header_row(rows)
    item_col = _column_index(header, "项目") or 1
    current_col = _column_index(header, current_header)
    ytd_col = _column_index(header, ytd_header)
    if not current_col and not ytd_col:
        raise StatementParseError(f"{path.name}: cannot find amount columns")

    current: Dict[str, Optional[float]] = {}
    ytd: Dict[str, Optional[float]] = {}
    lineage: Dict[str, SourceCell] = {}
    for row_idx, row in enumerate(rows[header_row_idx:], start=header_row_idx + 1):
        item = _normalize_item(_get_row_value(row, item_col))
        if not item:
            continue
        if current_col:
            value = _as_number(_get_row_value(row, current_col))
            current[item] = value
            lineage[f"{statement}.{quarter}.{item}.current"] = SourceCell(str(path), ws.title, row_idx, current_col, item, "current", value)
        if ytd_col:
            value = _as_number(_get_row_value(row, ytd_col))
            ytd[item] = value
            lineage[f"{statement}.{quarter}.{item}.ytd"] = SourceCell(str(path), ws.title, row_idx, ytd_col, item, "ytd", value)
    return current, ytd, lineage


def _parse_balance_sheet(ws: Any, *, path: Path, quarter: str) -> Tuple[Dict[str, Optional[float]], Dict[str, Optional[float]], Dict[str, SourceCell]]:
    rows = list(ws.iter_rows(values_only=True))
    header_row_idx, header = _find_header_row(rows)
    normalized = [_normalize_item(v) for v in header]

    asset_item_col = 1
    asset_ending_col = 3 if len(header) >= 3 else None
    right_item_col = None
    right_ending_col = None
    for idx, value in enumerate(normalized, start=1):
        if "负债" in value and "权益" in value:
            right_item_col = idx
        elif value == "期末余额" and idx > 4 and right_ending_col is None:
            right_ending_col = idx
    if not asset_ending_col or not right_item_col or not right_ending_col:
        raise StatementParseError(f"{path.name}: cannot identify balance-sheet blocks")

    assets: Dict[str, Optional[float]] = {}
    liab_equity: Dict[str, Optional[float]] = {}
    lineage: Dict[str, SourceCell] = {}
    for row_idx, row in enumerate(rows[header_row_idx:], start=header_row_idx + 1):
        asset_item = _normalize_item(_get_row_value(row, asset_item_col))
        if asset_item:
            value = _as_number(_get_row_value(row, asset_ending_col))
            assets[asset_item] = value
            lineage[f"balance_asset.{quarter}.{asset_item}.ending"] = SourceCell(str(path), ws.title, row_idx, asset_ending_col, asset_item, "ending", value)
        right_item = _normalize_item(_get_row_value(row, right_item_col))
        if right_item:
            value = _as_number(_get_row_value(row, right_ending_col))
            liab_equity[right_item] = value
            lineage[f"balance_liability_equity.{quarter}.{right_item}.ending"] = SourceCell(str(path), ws.title, row_idx, right_ending_col, right_item, "ending", value)
    return assets, liab_equity, lineage


def _get_row_value(row: Tuple[Any, ...], one_based_col: int) -> Any:
    idx = one_based_col - 1
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _validate_required_quarters(facts: ReportFacts) -> None:
    missing: List[str] = []
    for quarter in facts.quarters:
        if quarter not in facts.profit_current:
            missing.append(f"{quarter} 利润表")
        if quarter not in facts.cashflow_current:
            missing.append(f"{quarter} 现金流量表")
        if quarter not in facts.balance_assets_ending:
            missing.append(f"{quarter} 资产负债表")
    if missing:
        raise StatementParseError("Missing required statements: " + ", ".join(missing))
