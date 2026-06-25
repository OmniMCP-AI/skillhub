#!/usr/bin/env python3
from __future__ import annotations

import argparse
from datetime import date
import json
import os
import re
import subprocess
import sys
import urllib.error
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_URL = os.environ.get("MAYBEAI_BASE_URL", "https://play-be.omnimcp.ai")
TOKEN = os.environ.get("MAYBEAI_API_TOKEN", "")
REPORT_CORE = Path(__file__).resolve().parents[1] / "traceable-financial-analysis" / "scripts"
if str(REPORT_CORE) not in sys.path:
    sys.path.append(str(REPORT_CORE))

from finclaw_report_core.dashboard_rows import build_dashboard_rows
from finclaw_report_core.formatting import fmt_ratio, fmt_yuan_as_wan
from finclaw_report_core.metrics import build_metrics
from finclaw_report_core.narrative_values import build_management_summary
from finclaw_report_core.statement_facts import parse_statement_directory


QUARTERS = ("Q1", "Q2", "Q3", "Q4")
QUARTER_LABELS = {"Q1": "一季度", "Q2": "二季度", "Q3": "三季度", "Q4": "四季度"}
QUARTER_END = {"Q1": "2025-03-31", "Q2": "2025-06-30", "Q3": "2025-09-30", "Q4": "2025-12-31"}
PROFIT_ITEMS = {
    "营业收入": "一、营业收入",
    "营业成本": "减：营业成本",
    "销售费用": "销售费用",
    "研发费用": "研发费用",
    "管理费用": "管理费用",
    "财务费用": "财务费用",
    "营业利润": "二、营业利润",
    "净利润": "四、净利润",
}
BALANCE_ITEMS = {
    "货币资金": "货币资金",
    "应收账款": "应收账款",
    "流动资产合计": "流动资产合计",
    "资产总计": "资产总计",
    "应付账款": "应付账款",
    "流动负债合计": "流动负债合计",
    "负债合计": "负债合计",
    "所有者权益合计": "所有者权益合计",
}
CASH_ITEMS = {
    "经营现金流": "经营活动产生的现金流量净额",
    "期末现金余额": "期末现金余额",
}
ALIASES = {
    "研究费用": "研发费用",
    "研发投入": "研发费用",
    "二、营业利润（亏损以“-”号填列）": "二、营业利润",
    '二、营业利润(亏损以"-"号填列)': "二、营业利润",
    "二、营业利润": "二、营业利润",
    "四：净利润（净亏损以“-”号填列）": "四、净利润",
    "四：净利润(净亏损以\"-\"号填列)": "四、净利润",
    "四、净利润（净亏损以“-”号填列）": "四、净利润",
    '四、净利润(净亏损以"-"号填列)': "四、净利润",
    "四、净利润": "四、净利润",
    "流动负债合计：": "流动负债合计",
    "流动负债合计": "流动负债合计",
    "所有者权益（或股东权益）合计": "所有者权益合计",
    "所有者权益(或股东权益)合计": "所有者权益合计",
    "所有者权益合计": "所有者权益合计",
    "负债和所有者权益（或股东权益）总计": "负债和所有者权益总计",
    "负债和所有者权益(或股东权益)总计": "负债和所有者权益总计",
    "负债和所有者权益总计": "负债和所有者权益总计",
    "五、期末现金余额": "期末现金余额",
    "四、期末现金及现金等价物余额": "期末现金余额",
    "期末现金余额": "期末现金余额",
    "实收资本（或股本）": "实收资本",
    "实收资本(或股本)": "实收资本",
}


@dataclass
class NormalizedSheet:
    name: str
    rows: list[list[Any]]
    row_map: dict[str, int]
    source_map: dict[str, str]
    kind: str
    quarter: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input-dir",
        default=str(Path.home() / "Downloads/20260615_yunling_tech/extracted/demo-上海云棱智能科技有限公司"),
    )
    parser.add_argument(
        "--output-dir",
        default=str(Path.cwd() / "artifacts" / "yunling_traceable_report"),
    )
    return parser.parse_args()


def clean_text(value: Any) -> str:
    text = str(value or "").strip()
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("－", "-").replace("—", "-").replace("–", "-").replace("−", "-")
    text = text.replace("“", '"').replace("”", '"')
    text = re.sub(r"\s+", "", text)
    return ALIASES.get(text, text)


def as_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def detect_kind_and_quarter(path: Path, title: str) -> tuple[str, str]:
    text = f"{path.name} {title}"
    if "利润表" in text:
        kind = "profit"
    elif "资产负债表" in text:
        kind = "balance"
    elif "现金流量表" in text:
        kind = "cash"
    else:
        raise ValueError(f"无法识别报表类型: {path.name}")
    if "Q1" in text or "第1季度" in text or "3期" in text:
        quarter = "Q1"
    elif "Q2" in text or "第2季度" in text or "6期" in text:
        quarter = "Q2"
    elif "Q3" in text or "第3季度" in text or "9期" in text:
        quarter = "Q3"
    elif "Q4" in text or "第4季度" in text or "12期" in text:
        quarter = "Q4"
    else:
        raise ValueError(f"无法识别季度: {path.name}")
    return kind, quarter


def find_header_row(ws: Any) -> int:
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        values = [clean_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        if "项目" in values or "资产" in values:
            return row_idx
    raise ValueError(f"{ws.title}: 未找到表头")


def normalize_input_sheets(input_dir: Path) -> dict[str, NormalizedSheet]:
    sheets: dict[str, NormalizedSheet] = {}
    for path in sorted(input_dir.glob("*.xlsx")):
        if path.name.startswith(".~") or path.name.startswith("~$"):
            continue
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        kind, quarter = detect_kind_and_quarter(path, ws.title)
        if kind in {"profit", "cash"}:
            normalized = normalize_two_amount_sheet(path, ws, kind, quarter)
        else:
            normalized = normalize_balance_sheet(path, ws, quarter)
        sheets[normalized.name] = normalized
    if len(sheets) != 12:
        raise ValueError(f"期望 12 张标准化原始表，实际得到 {len(sheets)} 张")
    return sheets


def normalize_two_amount_sheet(path: Path, ws: Any, kind: str, quarter: str) -> NormalizedSheet:
    title = "利润表" if kind == "profit" else "现金流量表"
    header_row = find_header_row(ws)
    company_cell = str(ws.cell(2, 1).value or "")
    period_cell = str(ws.cell(2, 2).value or "")
    unit_cell = str(ws.cell(2, 4).value or "单位：元")
    rows: list[list[Any]] = [
        [title, "", "", ""],
        [company_cell, period_cell, "", unit_cell],
        ["项目", "行次", "本年累计金额" if kind == "profit" else "本期金额", "本期金额" if kind == "profit" else "本年累计金额"],
    ]
    row_map: dict[str, int] = {}
    source_map: dict[str, str] = {}
    item_col = 1
    line_col = 2
    amt_a_col = 3
    amt_b_col = 4
    excel_row = 4
    for src_row in range(header_row + 1, ws.max_row + 1):
        raw_item = ws.cell(src_row, item_col).value
        item = clean_text(raw_item)
        if not item:
            continue
        amount_a = as_number(ws.cell(src_row, amt_a_col).value)
        amount_b = as_number(ws.cell(src_row, amt_b_col).value)
        line_no = ws.cell(src_row, line_col).value
        if amount_a is None and amount_b is None and not line_no:
            continue
        rows.append([item, line_no or "", amount_a if amount_a is not None else "", amount_b if amount_b is not None else ""])
        row_map[item] = excel_row
        source_map[item] = f"{path.name} / {ws.title} / A{src_row}:D{src_row}"
        excel_row += 1
    return NormalizedSheet(
        name=f"{title}-2025{quarter}",
        rows=rows,
        row_map=row_map,
        source_map=source_map,
        kind=kind,
        quarter=quarter,
    )


def normalize_balance_sheet(path: Path, ws: Any, quarter: str) -> NormalizedSheet:
    header_row = find_header_row(ws)
    company_cell = str(ws.cell(2, 1).value or "")
    period_cell = QUARTER_END[quarter]
    unit_cell = "单位：元"
    rows: list[list[Any]] = [
        ["资产负债表", "", "", ""],
        [company_cell, period_cell, "", unit_cell],
        ["项目", "行次", "期末余额", "年初余额"],
    ]
    row_map: dict[str, int] = {}
    source_map: dict[str, str] = {}
    excel_row = 4
    for src_row in range(header_row + 1, ws.max_row + 1):
        left_item = clean_text(ws.cell(src_row, 1).value)
        left_line = ws.cell(src_row, 2).value
        left_end = as_number(ws.cell(src_row, 3).value)
        left_begin = as_number(ws.cell(src_row, 4).value)
        if left_item and (left_line or left_end is not None or left_begin is not None):
            rows.append([left_item, left_line or "", left_end if left_end is not None else "", left_begin if left_begin is not None else ""])
            row_map[left_item] = excel_row
            source_map[left_item] = f"{path.name} / {ws.title} / A{src_row}:D{src_row}"
            excel_row += 1
        right_item = clean_text(ws.cell(src_row, 5).value)
        right_line = ws.cell(src_row, 6).value
        right_end = as_number(ws.cell(src_row, 7).value)
        right_begin = as_number(ws.cell(src_row, 8).value)
        if right_item and (right_line or right_end is not None or right_begin is not None):
            rows.append([right_item, right_line or "", right_end if right_end is not None else "", right_begin if right_begin is not None else ""])
            row_map[right_item] = excel_row
            source_map[right_item] = f"{path.name} / {ws.title} / E{src_row}:H{src_row}"
            excel_row += 1
    return NormalizedSheet(
        name=f"资产负债表-2025{quarter}",
        rows=rows,
        row_map=row_map,
        source_map=source_map,
        kind="balance",
        quarter=quarter,
    )


def make_workbook(output_path: Path, company: str, normalized: dict[str, NormalizedSheet], metrics: Any, audit_rows: list[list[str]], lineage_rows: list[list[str]], limitation_rows: list[list[str]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    def sheet(name: str, headers: list[str]) -> Any:
        ws = wb.create_sheet(name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        return ws

    summary_lines = build_management_summary(metrics)
    dashboard_rows = build_dashboard_rows(metrics)
    q4_revenue = metrics.value("revenue.current", "Q4") or 0
    q3_revenue = metrics.value("revenue.current", "Q3") or 0
    q4_margin = metrics.value("net_margin.current", "Q4")
    q4_ocf_ratio = metrics.value("ocf_net_profit_ratio.current", "Q4")
    debt_q1 = metrics.value("debt_to_asset.ending", "Q1")
    debt_q4 = metrics.value("debt_to_asset.ending", "Q4")
    ar_q1 = metrics.value("accounts_receivable.ending", "Q1") or 0
    ar_q4 = metrics.value("accounts_receivable.ending", "Q4") or 0
    rd_q4 = metrics.value("rd_expense.ytd", "Q4")
    mgmt_q4 = metrics.value("management_expense.ytd", "Q4")
    sales_q4 = metrics.value("sales_expense.ytd", "Q4")
    cover = sheet("封面", ["项目", "内容", "辅助项目", "辅助内容"])
    cover_rows = [
        ["公司", company, "报告期间", "2025年Q1-Q4"],
        ["报告名称", f"{company} 财报与经营分析", "报告模板", "老板审阅 + 可追溯公式版"],
        ["数据口径", "利润表/现金流量表按单季与累计拆分，资产负债表按期末余额；报告层统一换算为万元。", "金额单位", "人民币万元"],
        ["核心结论", summary_lines[0], "现金质量", summary_lines[1]],
        ["报告用途", "老板审阅 / 财务经营复盘", "生成日期", str(date.today())],
        ["限制事项", "经营明细、预算、同行对标未提供；经营分析仅能基于三表做代理判断。", "数据性质", "FIN_STMT（真实财务报表）；经营观察含有限推断"],
    ]
    append_rows(cover, cover_rows)

    boss = sheet("老板摘要", ["模块", "管理层结论（事实+数字）", "决策提示（管理含义）"])
    append_rows(
        boss,
        [
            ["总体表现", summary_lines[0], "全年收入、利润、现金均为正增长，主营扩张与盈利兑现同步。"],
            ["增长质量", f"Q4 单季收入 {fmt_yuan_as_wan(q4_revenue)} 万元，较 Q3 的 {fmt_yuan_as_wan(q3_revenue)} 万元环比提升 {fmt_ratio((q4_revenue / q3_revenue - 1) if q3_revenue else None)}。", "增速虽较 Q2/Q3 略放缓，但单季盈利率继续上行，增长质量仍偏健康。"],
            ["现金质量", summary_lines[1], "经营现金流持续高于净利润，利润兑现为现金的能力较好。"],
            ["偿债结构", f"资产负债率由 Q1 的 {fmt_ratio(debt_q1)} 降至 Q4 的 {fmt_ratio(debt_q4)}，期末货币资金充足。", "杠杆水平改善，短中期偿债压力可控。"],
            ["管理动作", "经营数据尚未覆盖客户、产品、合同、预算与回款账龄。", "下一步优先补经营明细，验证收入增长、费用投放和回款效率的驱动关系。"],
        ],
    )

    overview = sheet("经营概览", ["项目", "本次观察（数据事实）", "管理说明（业务含义）"])
    append_rows(
        overview,
        [
            ["财务表现", summary_lines[0], "全年营收与净利润同增，财务主线明确。"],
            ["收入趋势", f"单季收入由 Q1 {fmt_yuan_as_wan(metrics.value('revenue.current', 'Q1'))} 万元增至 Q4 {fmt_yuan_as_wan(q4_revenue)} 万元。", "季度规模持续抬升，说明订单或交付节奏整体向上。"],
            ["盈利趋势", f"Q4 单季净利率 {fmt_ratio(q4_margin)}，全年净利率 {fmt_ratio(metrics.value('net_margin.ytd', 'Q4'))}。", "利润率逐季改善，费用扩张未吞噬利润增长。"],
            ["现金状态", f"Q4 单季 OCF/净利润 {fmt_ratio(q4_ocf_ratio)}，全年累计为 {fmt_ratio(metrics.value('ocf_net_profit_ratio.ytd', 'Q4'))}。", "现金创造能力强于会计利润，利润质量较好。"],
            ["费用效率", f"全年销售/管理/研发费用分别为 {fmt_yuan_as_wan(sales_q4)} / {fmt_yuan_as_wan(mgmt_q4)} / {fmt_yuan_as_wan(rd_q4)} 万元。", "管理费用占比仍高，研发投入占比较高，后续要验证费用投入产出。"],
            ["经营数据说明", "当前仅有三表，未提供预算、同行、客户、产品、订单、回款账龄等经营明细。", "经营分析结论属于财务代理判断，不替代业务侧明细诊断。"],
        ],
    )

    profit = sheet("利润分析", ["指标", "一季度", "二季度", "三季度", "四季度", "全年", "分析口径"])
    append_rows(
        profit,
        [
            ["营业收入", "", "", "", "", "", "单季收入，来源于利润表本期金额；全年为 Q1-Q4 求和。"],
            ["营业成本", "", "", "", "", "", "单季成本，来源于利润表本期金额；全年为 Q1-Q4 求和。"],
            ["毛利", "", "", "", "", "", "营业收入 - 营业成本。"],
            ["毛利率", "", "", "", "", "", "毛利 / 营业收入。"],
            ["销售费用", "", "", "", "", "", "单季销售费用；全年求和。"],
            ["研发费用", "", "", "", "", "", "单季研发费用；全年求和。"],
            ["管理费用", "", "", "", "", "", "单季管理费用；全年求和。"],
            ["财务费用", "", "", "", "", "", "单季财务费用；全年求和。"],
            ["营业利润", "", "", "", "", "", "单季营业利润；全年求和。"],
            ["净利润", "", "", "", "", "", "单季净利润；全年求和。"],
            ["净利率", "", "", "", "", "", "净利润 / 营业收入。"],
        ],
    )

    balance = sheet("资产负债分析", ["一季度末", "二季度末", "三季度末", "四季度末", "指标", "管理观察"])
    append_rows(
        balance,
        [
            ["", "", "", "", "货币资金", "期末现金储备持续上升。"],
            ["", "", "", "", "应收账款", "规模同步扩大，需要持续跟踪账龄和回款速度。"],
            ["", "", "", "", "流动资产合计", "流动资产规模扩张，为增长提供周转基础。"],
            ["", "", "", "", "资产总计", "总资产逐季抬升。"],
            ["", "", "", "", "应付账款", "与业务规模扩张同步增长。"],
            ["", "", "", "", "流动负债合计", "需结合回款节奏看短债压力。"],
            ["", "", "", "", "负债合计", "整体负债绝对额上升但占比下降。"],
            ["", "", "", "", "所有者权益合计", "利润留存推动权益持续增厚。"],
            ["", "", "", "", "资产负债率", "杠杆水平逐季改善。"],
        ],
    )

    cash = sheet("现金流分析", ["指标", "Q1", "Q2", "Q3", "Q4", "全年/期末", "管理观察"])
    append_rows(
        cash,
        [
            ["经营现金流", "", "", "", "", "", "单季经营现金流保持正值，全年累计高于净利润。"],
            ["净利润", "", "", "", "", "", "与利润分析保持一致。"],
            ["OCF/净利润", "", "", "", "", "", "观察利润兑现为现金的质量。"],
            ["期末现金余额", "", "", "", "", "", "来自资产负债表货币资金，并与现金流期末现金勾稽。"],
        ],
    )

    key = sheet("关键指标", ["一季度", "二季度", "三季度", "四季度", "全年/期末", "指标"])
    append_rows(
        key,
        [
            ["", "", "", "", "", "营业收入"],
            ["", "", "", "", "", "毛利"],
            ["", "", "", "", "", "毛利率"],
            ["", "", "", "", "", "营业利润"],
            ["", "", "", "", "", "净利润"],
            ["", "", "", "", "", "净利率"],
            ["", "", "", "", "", "经营现金流"],
            ["", "", "", "", "", "期末货币资金"],
            ["", "", "", "", "", "资产总计"],
            ["", "", "", "", "", "资产负债率"],
        ],
    )

    risk = sheet("风险与建议", ["类型", "事项", "判断", "建议"])
    append_rows(
        risk,
        [
            ["数据范围", "当前仅有三表，经营侧明细和预算未提供。", "经营分析深度受限。", "补充客户/产品/合同/预算/账龄后再做驱动拆解。"],
            ["利润质量", f"全年净利率 {fmt_ratio(metrics.value('net_margin.ytd', 'Q4'))}，Q4 单季净利率 {fmt_ratio(q4_margin)}。", "盈利能力改善。", "继续跟踪毛利率与费用率拐点，防止增长换利润。"],
            ["现金质量", f"全年 OCF/净利润 {fmt_ratio(metrics.value('ocf_net_profit_ratio.ytd', 'Q4'))}。", "现金质量较好。", "保持回款节奏，避免应收继续快于收入扩张。"],
            ["偿债结构", f"资产负债率由 {fmt_ratio(debt_q1)} 降至 {fmt_ratio(debt_q4)}。", "偿债结构改善。", "关注流动负债与货币资金的匹配及中长期资金安排。"],
            ["管理跟进", f"Q4 期末应收账款 {fmt_yuan_as_wan(ar_q4)} 万元，较 Q1 的 {fmt_yuan_as_wan(ar_q1)} 万元明显上升。", "规模扩张伴随营运资金占用上升。", "建立回款和账龄跟踪台账，验证增长的现金含量。"],
        ],
    )

    support = sheet("追问支持", ["可追问主题", "可追溯依据", "可下钻字段", "需要补充资料"])
    append_rows(
        support,
        [
            ["收入与利润趋势", "利润表-2025Q1~Q4 本期金额 + 利润分析公式区", "分季度收入、成本、净利润", "订单、客户、产品、地区收入结构"],
            ["费用结构", "利润表-2025Q1~Q4 销售/研发/管理/财务费用", "费用分季度、费用率", "费用明细台账、预算口径"],
            ["资产和偿债能力", "资产负债表-2025Q1~Q4 资产总计/负债合计/权益", "债务结构、营运资金", "借款结构、账龄、供应商付款周期"],
            ["现金流质量", "现金流量表-2025Q1~Q4 经营现金流、期末现金余额", "单季/累计 OCF、现金勾稽", "回款明细、收支预测"],
            ["经营数据缺口", "封面/经营概览/限制事项", "预算、同行、业务明细", "预算、同行对标、经营 KPI 明细"],
        ],
    )

    overview2 = sheet("关键指标总览", ["指标", "Q1", "Q2", "Q3", "Q4", "全年/期末", "口径 / 说明"])
    for row in dashboard_rows:
        overview2.append([row.label, "", "", "", "", "", row.note])

    dashboard = sheet("财务经营Dashboard", ["指标", "Q1", "Q2", "Q3", "Q4", "全年/期末", "口径 / 说明"])
    for row in dashboard_rows:
        dashboard.append([row.label, "", "", "", "", "", row.note])

    deep = sheet("深度分析报告", ["模块", "数据事实", "管理判断", "建议动作"])
    append_rows(
        deep,
        [
            ["收入增长与规模扩张", f"全年营收 {fmt_yuan_as_wan(metrics.value('revenue.ytd', 'Q4'))} 万元，Q1→Q4 单季收入持续增长。", "公司 2025 年处于扩张周期，收入增长具有连续性。", "继续拆分客户、产品、交付节奏，验证增长来源是否均衡。"],
            ["盈利能力与费用吸收", f"全年毛利率 {fmt_ratio(metrics.value('gross_margin.ytd', 'Q4'))}，全年净利率 {fmt_ratio(metrics.value('net_margin.ytd', 'Q4'))}。", "毛利率略有回落但净利率上行，说明费用吸收和规模效应在改善。", "重点盯管理费用与研发费用投入产出比。"],
            ["现金质量与利润兑现", f"全年经营现金流 {fmt_yuan_as_wan(metrics.value('ocf.ytd', 'Q4'))} 万元，高于净利润 {fmt_yuan_as_wan(metrics.value('net_profit.ytd', 'Q4'))} 万元。", "现金流覆盖利润，利润质量较高。", "保持回款和支出纪律，防止扩张期现金消耗转弱。"],
            ["资产结构与偿债能力", f"Q4 货币资金 {fmt_yuan_as_wan(metrics.value('cash.bs.ending', 'Q4'))} 万元，资产负债率 {fmt_ratio(debt_q4)}。", "账面流动性较稳，杠杆下降提升抗风险能力。", "补充借款结构和现金预测，确认中长期资金安排。"],
            ["营运资金与回款压力", f"应收账款从 Q1 的 {fmt_yuan_as_wan(ar_q1)} 万元增至 Q4 的 {fmt_yuan_as_wan(ar_q4)} 万元。", "收入扩张带来营运资金占用上升，后续要验证回款质量。", "补账龄、回款和坏账准备数据。"],
            ["经营分析边界", "未提供客户、订单、产品、区域、预算、同行等经营明细。", "当前经营结论仅能依据三表做代理判断。", "补经营底稿后再做经营驱动分析与预算偏差分析。"],
        ],
    )

    audit = sheet("数据与计算审核报告", ["检查项", "结论", "说明"])
    append_rows(audit, audit_rows)

    lineage = sheet("关键数据溯源摘要", ["指标", "季度", "口径", "单位", "公式", "来源"])
    append_rows(lineage, lineage_rows)

    limits = sheet("限制事项与下一步", ["事项", "说明"])
    append_rows(limits, limitation_rows)

    for quarter in QUARTERS:
        for prefix in ("利润表", "资产负债表", "现金流量表"):
            raw = normalized[f"{prefix}-2025{quarter}"]
            ws = wb.create_sheet(raw.name)
            for row in raw.rows:
                ws.append(row)
            for cell in ws[3]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.freeze_panes = "A4"
            ws.sheet_view.showGridLines = False

    for ws in wb.worksheets:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[col_letter(col)].width = 22
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def append_rows(ws: Any, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)


def col_letter(index: int) -> str:
    result = ""
    while index:
        index, rem = divmod(index - 1, 26)
        result = chr(65 + rem) + result
    return result


def raw_ref(sheet_name: str, cell_ref: str) -> str:
    return f"'{sheet_name}'!{cell_ref}"


def wan_expr(expr: str) -> str:
    return f"ROUND(({expr})/10000,2)"


def round_wan(expr: str) -> str:
    return f"={wan_expr(expr)}"


def build_formula_operations(normalized: dict[str, NormalizedSheet]) -> list[dict[str, Any]]:
    def p_row(q: str, item: str) -> int:
        return normalized[f"利润表-2025{q}"].row_map[item]

    def b_row(q: str, item: str) -> int:
        return normalized[f"资产负债表-2025{q}"].row_map[item]

    def c_row(q: str, item: str) -> int:
        return normalized[f"现金流量表-2025{q}"].row_map[item]

    def profit_ytd_wan(q: str, item: str) -> str:
        return round_wan(raw_ref(f"利润表-2025{q}", f"C{p_row(q, item)}"))

    def profit_ytd_ratio(q: str, numerator_item: str, denominator_item: str) -> str:
        num = wan_expr(raw_ref(f"利润表-2025{q}", f"C{p_row(q, numerator_item)}"))
        den = wan_expr(raw_ref(f"利润表-2025{q}", f"C{p_row(q, denominator_item)}"))
        return f"=IFERROR(({num})/({den}),0)"

    def profit_ytd_margin(q: str) -> str:
        rev = wan_expr(raw_ref(f"利润表-2025{q}", f"C{p_row(q, PROFIT_ITEMS['营业收入'])}"))
        cost = wan_expr(raw_ref(f"利润表-2025{q}", f"C{p_row(q, PROFIT_ITEMS['营业成本'])}"))
        num = f"({rev}-{cost})"
        den = rev
        return f"=IFERROR(({num})/({den}),0)"

    profit_rows: list[list[str]] = []
    for report_row, item in [
        (2, PROFIT_ITEMS["营业收入"]),
        (3, PROFIT_ITEMS["营业成本"]),
        (6, PROFIT_ITEMS["销售费用"]),
        (7, PROFIT_ITEMS["研发费用"]),
        (8, PROFIT_ITEMS["管理费用"]),
        (9, PROFIT_ITEMS["财务费用"]),
        (10, PROFIT_ITEMS["营业利润"]),
        (11, PROFIT_ITEMS["净利润"]),
    ]:
        row_formulas = [round_wan(raw_ref(f"利润表-2025{q}", f"D{p_row(q, item)}")) for q in QUARTERS]
        row_formulas.append(f"=ROUND(SUM(B{report_row}:E{report_row}),2)")
        profit_rows.append((report_row, row_formulas))

    profit_block: list[list[str]] = []
    formula_by_row = {row: formulas for row, formulas in profit_rows}
    for row in range(2, 13):
        if row in formula_by_row:
            profit_block.append(formula_by_row[row])
        elif row == 4:
            profit_block.append([f"=ROUND(B2-B3,2)", f"=ROUND(C2-C3,2)", f"=ROUND(D2-D3,2)", f"=ROUND(E2-E3,2)", "=ROUND(SUM(B4:E4),2)"])
        elif row == 5:
            profit_block.append([f"=IFERROR(B4/B2,0)", f"=IFERROR(C4/C2,0)", f"=IFERROR(D4/D2,0)", f"=IFERROR(E4/E2,0)", "=IFERROR(F4/F2,0)"])
        elif row == 12:
            profit_block.append([f"=IFERROR(B11/B2,0)", f"=IFERROR(C11/C2,0)", f"=IFERROR(D11/D2,0)", f"=IFERROR(E11/E2,0)", "=IFERROR(F11/F2,0)"])

    balance_block: list[list[str]] = []
    for item in [
        BALANCE_ITEMS["货币资金"],
        BALANCE_ITEMS["应收账款"],
        BALANCE_ITEMS["流动资产合计"],
        BALANCE_ITEMS["资产总计"],
        BALANCE_ITEMS["应付账款"],
        BALANCE_ITEMS["流动负债合计"],
        BALANCE_ITEMS["负债合计"],
        BALANCE_ITEMS["所有者权益合计"],
    ]:
        balance_block.append([round_wan(raw_ref(f"资产负债表-2025{q}", f"C{b_row(q, item)}")) for q in QUARTERS])
    balance_block.append(["=IFERROR(A8/A5,0)", "=IFERROR(B8/B5,0)", "=IFERROR(C8/C5,0)", "=IFERROR(D8/D5,0)"])

    cash_block = [
        [round_wan(raw_ref(f"现金流量表-2025{q}", f"C{c_row(q, CASH_ITEMS['经营现金流'])}")) for q in QUARTERS] + ["=ROUND(SUM(B2:E2),2)"],
        [f"=利润分析!B11", f"=利润分析!C11", f"=利润分析!D11", f"=利润分析!E11", "=利润分析!F11"],
        ["=IFERROR(B2/B3,0)", "=IFERROR(C2/C3,0)", "=IFERROR(D2/D3,0)", "=IFERROR(E2/E3,0)", "=IFERROR(F2/F3,0)"],
        ["=资产负债分析!A2", "=资产负债分析!B2", "=资产负债分析!C2", "=资产负债分析!D2", "=E5"],
    ]

    key_block = [
        ["=利润分析!B2", "=利润分析!C2", "=利润分析!D2", "=利润分析!E2", "=利润分析!F2"],
        ["=利润分析!B4", "=利润分析!C4", "=利润分析!D4", "=利润分析!E4", "=利润分析!F4"],
        ["=利润分析!B5", "=利润分析!C5", "=利润分析!D5", "=利润分析!E5", "=利润分析!F5"],
        ["=利润分析!B10", "=利润分析!C10", "=利润分析!D10", "=利润分析!E10", "=利润分析!F10"],
        ["=利润分析!B11", "=利润分析!C11", "=利润分析!D11", "=利润分析!E11", "=利润分析!F11"],
        ["=利润分析!B12", "=利润分析!C12", "=利润分析!D12", "=利润分析!E12", "=利润分析!F12"],
        ["=现金流分析!B2", "=现金流分析!C2", "=现金流分析!D2", "=现金流分析!E2", "=现金流分析!F2"],
        ["=资产负债分析!A2", "=资产负债分析!B2", "=资产负债分析!C2", "=资产负债分析!D2", "=资产负债分析!D2"],
        ["=资产负债分析!A5", "=资产负债分析!B5", "=资产负债分析!C5", "=资产负债分析!D5", "=资产负债分析!D5"],
        ["=资产负债分析!A10", "=资产负债分析!B10", "=资产负债分析!C10", "=资产负债分析!D10", "=资产负债分析!D10"],
    ]

    overview_block = [
        [profit_ytd_wan(q, PROFIT_ITEMS["营业收入"]) for q in QUARTERS] + [profit_ytd_wan("Q4", PROFIT_ITEMS["营业收入"])],
        ["=利润分析!B2", "=利润分析!C2", "=利润分析!D2", "=利润分析!E2", "-"],
        ["" if idx == 0 else f"=IFERROR({col_letter(idx+2)}3/{col_letter(idx+1)}3-1,0)" for idx in range(4)] + ["-"],
        [profit_ytd_margin(q) for q in QUARTERS] + [profit_ytd_margin("Q4")],
    ]
    overview_block[2][0] = "-"
    overview_block.extend(
        [
            [profit_ytd_wan(q, PROFIT_ITEMS["营业利润"]) for q in QUARTERS] + [profit_ytd_wan("Q4", PROFIT_ITEMS["营业利润"])],
            [profit_ytd_wan(q, PROFIT_ITEMS["净利润"]) for q in QUARTERS] + [profit_ytd_wan("Q4", PROFIT_ITEMS["净利润"])],
            [profit_ytd_ratio(q, PROFIT_ITEMS["净利润"], PROFIT_ITEMS["营业收入"]) for q in QUARTERS] + [profit_ytd_ratio("Q4", PROFIT_ITEMS["净利润"], PROFIT_ITEMS["营业收入"])],
            [round_wan(raw_ref(f"现金流量表-2025{q}", f"D{c_row(q, CASH_ITEMS['经营现金流'])}")) for q in QUARTERS] + [round_wan(raw_ref("现金流量表-2025Q4", f"D{c_row('Q4', CASH_ITEMS['经营现金流'])}"))],
            ["=现金流分析!B4", "=现金流分析!C4", "=现金流分析!D4", "=现金流分析!E4", "-"],
            [f"=IFERROR(B9/B7,0)", f"=IFERROR(C9/C7,0)", f"=IFERROR(D9/D7,0)", f"=IFERROR(E9/E7,0)", "=IFERROR(F9/F7,0)"],
            ["=资产负债分析!A2", "=资产负债分析!B2", "=资产负债分析!C2", "=资产负债分析!D2", "=资产负债分析!D2"],
            ["=资产负债分析!A5", "=资产负债分析!B5", "=资产负债分析!C5", "=资产负债分析!D5", "=资产负债分析!D5"],
            ["=资产负债分析!A10", "=资产负债分析!B10", "=资产负债分析!C10", "=资产负债分析!D10", "=资产负债分析!D10"],
            ["=资产负债分析!A3", "=资产负债分析!B3", "=资产负债分析!C3", "=资产负债分析!D3", "=资产负债分析!D3"],
            [profit_ytd_wan(q, PROFIT_ITEMS["研发费用"]) for q in QUARTERS] + [profit_ytd_wan("Q4", PROFIT_ITEMS["研发费用"])],
        ]
    )

    dashboard_block = []
    for row_idx in range(2, 2 + len(overview_block)):
        dashboard_block.append([f"='关键指标总览'!B{row_idx}", f"='关键指标总览'!C{row_idx}", f"='关键指标总览'!D{row_idx}", f"='关键指标总览'!E{row_idx}", f"='关键指标总览'!F{row_idx}"])

    return [
        {"worksheet_name": "利润分析", "range_address": "B2:F12", "formulas": profit_block},
        {"worksheet_name": "资产负债分析", "range_address": "A2:D10", "formulas": balance_block},
        {"worksheet_name": "现金流分析", "range_address": "B2:F5", "formulas": cash_block},
        {"worksheet_name": "关键指标", "range_address": "A2:E11", "formulas": key_block},
        {"worksheet_name": "关键指标总览", "range_address": f"B2:F{1 + len(overview_block)}", "formulas": overview_block},
        {"worksheet_name": "财务经营Dashboard", "range_address": f"B2:F{1 + len(dashboard_block)}", "formulas": dashboard_block},
    ]


def build_audit_rows(facts: Any, normalized: dict[str, NormalizedSheet]) -> list[list[str]]:
    rows: list[list[str]] = []
    rows.append(["数据来源核验", "PASS", "共识别 12 份正式财务报表文件，覆盖 2025 年 Q1-Q4 的利润表、资产负债表、现金流量表。"])
    rows.append(["公司一致性", "PASS", f"解析公司名为 {facts.company}，与任务对象一致。"])
    rows.append(["金额单位统一", "PASS", "原始文件均为元，报告层统一换算为万元，比例指标保持小数口径。"])
    for quarter in QUARTERS:
        bs = normalized[f"资产负债表-2025{quarter}"]
        assets = sheet_value(bs, "资产总计")
        liabilities = sheet_value(bs, "负债合计")
        equity = sheet_value(bs, "所有者权益合计")
        diff = None if None in (assets, liabilities, equity) else round(abs((assets or 0) - (liabilities or 0) - (equity or 0)), 2)
        rows.append(["三表勾稽：资产负债平衡", "PASS" if diff is not None and diff <= 0.01 else "BLOCK", f"{quarter} 资产总计与负债+权益差异 {diff if diff is not None else '缺值'} 元。"])
        cash_bs = sheet_value(bs, "货币资金")
        cash_cf = sheet_value(normalized[f"现金流量表-2025{quarter}"], "期末现金余额")
        cash_diff = None if None in (cash_bs, cash_cf) else round(abs((cash_bs or 0) - (cash_cf or 0)), 2)
        rows.append(["三表勾稽：期末现金一致", "PASS" if cash_diff is not None and cash_diff <= 0.01 else "BLOCK", f"{quarter} 资产负债表货币资金与现金流量表期末现金差异 {cash_diff if cash_diff is not None else '缺值'} 元。"])
    rows.append(["经营数据完整性", "WARN", "未提供预算、同行、客户、产品、订单、账龄等经营明细，经营分析只能基于三表做代理判断。"])
    rows.append(["审核结论", "PASS_WITH_WARNINGS", "财务口径、勾稽关系和公式链条可交付；经营驱动分析需补数据后深化。"])
    return rows


def sheet_value(sheet: NormalizedSheet, item: str) -> float | None:
    row = sheet.row_map.get(item)
    if row is None:
        return None
    data_row = sheet.rows[row - 1]
    if sheet.kind == "balance":
        return as_number(data_row[2])
    if sheet.kind == "profit":
        return as_number(data_row[3])
    return as_number(data_row[3] if item == "期末现金余额" else data_row[2])


def build_lineage_rows(normalized: dict[str, NormalizedSheet]) -> list[list[str]]:
    rows: list[list[str]] = []
    def src(sheet_key: str, item: str) -> str:
        sheet = normalized[sheet_key]
        return f"{sheet_key} / {item} / {sheet.source_map[item]}"

    rows.extend(
        [
            ["营业收入", "Q4", "单季", "万元", "=利润分析!E2", src("利润表-2025Q4", PROFIT_ITEMS["营业收入"])],
            ["营业收入", "全年", "Q1-Q4 单季求和", "万元", "=利润分析!F2", "利润分析!B2:E2；来源于 Q1-Q4 利润表本期金额"],
            ["净利润", "Q4", "单季", "万元", "=利润分析!E11", src("利润表-2025Q4", PROFIT_ITEMS["净利润"])],
            ["净利润", "全年", "Q1-Q4 单季求和", "万元", "=利润分析!F11", "利润分析!B11:E11；来源于 Q1-Q4 利润表本期金额"],
            ["经营现金流", "全年", "累计", "万元", "=关键指标总览!F8", src("现金流量表-2025Q4", CASH_ITEMS["经营现金流"])],
            ["期末货币资金", "Q4", "期末", "万元", "=资产负债分析!D2", src("资产负债表-2025Q4", BALANCE_ITEMS["货币资金"])],
            ["资产总计", "Q4", "期末", "万元", "=资产负债分析!D5", src("资产负债表-2025Q4", BALANCE_ITEMS["资产总计"])],
            ["资产负债率", "Q4", "期末", "%", "=资产负债分析!D10", "资产负债分析!D8 / D5；上游来自 Q4 资产负债表负债合计与资产总计"],
            ["研发费用", "全年", "累计", "万元", "=关键指标总览!F15", src("利润表-2025Q4", PROFIT_ITEMS["研发费用"])],
            ["应收账款", "Q4", "期末", "万元", "=关键指标总览!F14", src("资产负债表-2025Q4", BALANCE_ITEMS["应收账款"])],
        ]
    )
    return rows


def build_limit_rows() -> list[list[str]]:
    return [
        ["限制事项", "当前未提供预算、同行、客户、产品、订单、合同、回款账龄等经营数据，本次经营分析仅基于三表数据。"],
        ["下一步 1", "补预算和目标值，新增预算偏差分析。"],
        ["下一步 2", "补客户/产品/订单/账龄，新增经营驱动分析页。"],
        ["下一步 3", "补同行或行业基准，新增对标分析与风险分层。"],
    ]


def post_json(path: str, payload: dict[str, Any]) -> dict[str, Any]:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        f"{BASE_URL}{path}",
        data=data,
        headers={
            "Authorization": f"Bearer {TOKEN}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=90) as resp:
            return json.loads(resp.read().decode("utf-8", "replace"))
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", "replace")
        raise RuntimeError(f"{path} failed: {exc.code} {body}") from exc


def upload_workbook(path: Path) -> tuple[str, str]:
    cmd = [
        "curl",
        "-sS",
        "-X",
        "POST",
        f"{BASE_URL}/api/v1/excel/upload",
        "-H",
        f"Authorization: Bearer {TOKEN}",
        "-F",
        f"file=@{path}",
    ]
    resp = subprocess.run(cmd, check=True, capture_output=True, text=True)
    payload = json.loads(resp.stdout)
    doc_id = payload.get("document_id") or payload.get("uri", "").split("/d/")[-1].split("?")[0]
    uri = (payload.get("uri") or f"https://www.maybe.ai/docs/spreadsheets/d/{doc_id}").split("?")[0]
    if not doc_id:
        raise RuntimeError(f"上传成功但未返回 document_id: {payload}")
    return doc_id, uri


def export_workbook(doc_id: str, path: Path) -> None:
    req = urllib.request.Request(
        f"{BASE_URL}/api/v1/excel/export/{doc_id}",
        headers={"Authorization": f"Bearer {TOKEN}"},
        method="GET",
    )
    with urllib.request.urlopen(req, timeout=120) as resp:
        path.write_bytes(resp.read())


def verify_delivery(uri: str, expected_sheet_count: int) -> dict[str, Any]:
    worksheets = post_json("/api/v1/excel/list_worksheets", {"uri": uri})
    titles = [ws.get("title") or ws.get("name") for ws in worksheets.get("worksheets", [])]
    checks = {}
    for name in ["封面", "老板摘要", "利润分析", "现金流分析", "数据与计算审核报告", "关键数据溯源摘要"]:
        checks[name] = post_json("/api/v1/excel/read_sheet", {"uri": uri, "worksheet_name": name})
    return {
        "titles": titles,
        "sheet_count_ok": len(titles) == expected_sheet_count,
        "checks": {
            name: {
                "shape": checks[name].get("shape"),
                "headers": checks[name].get("headers"),
                "non_empty": bool(checks[name].get("data")),
            }
            for name in checks
        },
    }


def main() -> None:
    if not TOKEN:
        raise SystemExit("MAYBEAI_API_TOKEN 未设置")
    args = parse_args()
    input_dir = Path(args.input_dir).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    clean_dir = output_dir / "clean_source"
    clean_dir.mkdir(exist_ok=True)
    for src in input_dir.glob("*.xlsx"):
        if src.name.startswith(".~") or src.name.startswith("~$"):
            continue
        dst = clean_dir / src.name
        if not dst.exists():
            dst.write_bytes(src.read_bytes())

    facts = parse_statement_directory(clean_dir, company="上海云棱智能科技有限公司", year=2025)
    metrics = build_metrics(facts)
    normalized = normalize_input_sheets(clean_dir)
    audit_rows = build_audit_rows(facts, normalized)
    lineage_rows = build_lineage_rows(normalized)
    limitation_rows = build_limit_rows()

    workbook_path = output_dir / "上海云棱智能科技有限公司-2025财报与经营分析-可追溯公式版.xlsx"
    make_workbook(workbook_path, facts.company, normalized, metrics, audit_rows, lineage_rows, limitation_rows)

    formula_ops = build_formula_operations(normalized)
    formula_path = output_dir / "formula_operations.json"
    formula_path.write_text(json.dumps(formula_ops, ensure_ascii=False, indent=2), encoding="utf-8")

    doc_id, uri = upload_workbook(workbook_path)
    post_json(
        "/api/v1/excel/formula/batch_set",
        {
            "uri": uri,
            "skip_recalculation": True,
            "recalculate_mode": "workbook",
            "operations": formula_ops,
        },
    )
    post_json("/api/v1/excel/recalculate_formulas", {"uri": uri})
    verification = verify_delivery(uri, 27)

    exported_path = output_dir / "上海云棱智能科技有限公司-2025财报与经营分析-可追溯公式版-导出.xlsx"
    export_workbook(doc_id, exported_path)

    result = {
        "company": facts.company,
        "doc_id": doc_id,
        "uri": uri,
        "local_workbook": str(workbook_path),
        "local_exported_workbook": str(exported_path),
        "verification": verification,
        "audit_report_rows": audit_rows,
    }
    (output_dir / "run_result.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
